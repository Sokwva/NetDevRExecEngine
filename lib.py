import logging
import os
import pathlib
from typing import Any

import netmiko
import netmiko.exceptions
import openpyxl
import openpyxl.worksheet.worksheet
from netmiko import NetmikoTimeoutException


class DeviceExcel:

    def __init__(self, excel_file, enable_log=True):
        self.main_Log_handler = self.set_log(enable_log)
        self.target_list = self.__reade_excel_to_dicts(excel_file)
        pass

    @staticmethod
    def set_log(enable_log: bool = True):
        """

        :param enable_log:
        :return: logging
        """
        main_log_handler = logging.getLogger(__name__)
        main_log_handler.setLevel(logging.INFO)
        log_file_handler = logging.FileHandler(filename="exec.log", mode="a")
        log_formatter = logging.Formatter(
            '%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s')
        log_stream_handler = logging.StreamHandler()
        log_file_handler.setFormatter(log_formatter)
        log_stream_handler.setFormatter(log_formatter)
        main_log_handler.addHandler(log_stream_handler)
        if enable_log:
            main_log_handler.addHandler(log_file_handler)
        return main_log_handler

    def __reade_excel_to_dicts(self, excel_file) -> list[dict[str:str]]:
        self.main_Log_handler.info("Init Excel Worksheet.")
        ws: openpyxl.worksheet.worksheet.Worksheet = openpyxl.load_workbook(
            excel_file, read_only=True).active
        # print(workFilePath, wb.sheetnames)
        # 行数
        i = 0
        header = []
        target_list = []
        self.main_Log_handler.info("Make targetList.")
        for row in ws.iter_rows():
            exe_target = {}
            # 列数
            n = 0
            for cell in row:
                if i == 0:
                    header.append(cell.value)
                else:
                    if cell.value is None:
                        break
                    exe_target[header[n]] = cell.value
                n += 1
            if i == 0:
                i += 1
                continue
            else:
                i += 1
                if exe_target == {}:
                    continue
                target_list.append(exe_target)
        return target_list

    def run(self):
        self.main_Log_handler.info("Make Fail_list.")
        fail_list: list[dict[str, NetmikoTimeoutException | Any]] = []
        for target in self.target_list:
            if target["ip"] is None:
                continue
            self.main_Log_handler.info("Make connect_info: " + target["ip"])
            connect_info = {
                "ip": target["ip"],
                "username": target["username"],
                "password": target["password"],
                "session_log": target["session_log"],
                "encoding": "utf-8" if target["encoding"] is None else target["encoding"],
                "device_type": target["device_type"],
            }

            if not pathlib.Path(connect_info.get(
                    "session_log")).exists():
                pathlib.Path(connect_info.get(
                    "session_log")).parent.mkdir(parents=True)
                self.main_Log_handler.info("Create session log file")

            try:
                conn = netmiko.ConnectHandler(**connect_info)
                self.main_Log_handler.info("Connecting " + target["ip"])
            except netmiko.exceptions.SSHException as e:
                self.main_Log_handler.warning("Exec " + target["ip"] + "failed.")
                fail = {
                    "ip": connect_info["ip"],
                    "error_msg": e
                }
                fail_list.append(fail)
                continue
            self.main_Log_handler.info("Exec " + target["ip"])
            conn.send_config_from_file(target["cfg_file"])
        return fail_list

    def run_as_save_result(self, result_csv_file):
        result = self.run()
        if len(result) == 0:
            return
        pathlib.Path(result_csv_file).parent.mkdir(parents=True, exist_ok=True)
        import csv
        with open(result_csv_file, mode="w", newline="") as f:
            write = csv.DictWriter(f, fieldnames=result[0].keys())
            write.writeheader()
            write.writerows(result)

        os.startfile(pathlib.Path(result_csv_file))


class MultiDevExcel(DeviceExcel):
    def __init__(self, excel_file):
        super().__init__(excel_file)
