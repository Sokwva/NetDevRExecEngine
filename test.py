from openpyxl import Workbook, load_workbook
import os
import openpyxl.worksheet.worksheet
import paramiko
import time

basePath = os.path.abspath(os.path.dirname(__file__))
workFilePath = os.path.join(basePath, "test.xlsx")
wb: Workbook = load_workbook(workFilePath, read_only=True)
ws: openpyxl.worksheet.worksheet.Worksheet = wb.active


def main():
    # print(workFilePath, wb.sheetnames)
    # 行数
    i = 0
    header = []
    for row in ws.iter_rows():
        exeTarget = {}
        # 列数
        n = 0
        for cell in row:
            if i == 0:
                header.append(cell.value)
            else:
                exeTarget[header[n]] = cell.value
            n += 1
        # print(i, exeTarget)
        if i != 0:
            connect(exeTarget['IP'], exeTarget['User'],
                    exeTarget['Password'], exeTarget['CMD'], exeTarget['LogFile'])
        i += 1
    # print(header)


def connect(ip, user, password, cmds, logName="exec.log"):
    if ip is None or user is None or password is None or cmds is None:
        print("param is none")
        return
    if type(cmds) != list:
        print("cmds param is not list")
        x = [cmds]
        cmds = x

    logFile = open(os.path.join(basePath, "logs\\"+logName), mode="w")
    sshHandler = paramiko.SSHClient()
    sshHandler.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    try:
        sshHandler.connect(ip, 22, user, password, timeout=10,
                           allow_agent=False, look_for_keys=False)
        sshShell: paramiko.Channel = sshHandler.invoke_shell()
        print(sshShell.recv(999))
        for cmd in cmds:
            # print("cmd: ", cmd)
            sshShell.sendall(cmd+"\n")
            time.sleep(2)
        strRet = str(sshShell.recv(99999), encoding="utf-8")
        # print(strRet)
        logFile.writelines(strRet)
        sshHandler.close()
    except:
        print("connect host %s faild", ip)


if __name__ == "__main__":
    main()
